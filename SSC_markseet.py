import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills


def Metric_marksheet(wb,filename):
    """

    :param wb:
    :param filename:
    :return:
    """
    try:
        ws = wb.create_sheet('10th Marksheet')
        ws.sheet_properties.tabColor = '00000000'
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["A"].width = 5
        ws["B2"].value = "MARKSHEET"
        ws.cell(row=2,column=2).fill=PatternFill(fill_type=fills.FILL_SOLID,start_color='1B2631')
        ws.cell(row=2,column=2).font=Font(color='FFFFFFFF',bold=True,size=12,name='Ariel')
        ws.cell(row=2,column=2).alignment=Alignment(horizontal='center')
        sub = ['Physics','Chemistry','Biology','Mathematics','English','Hindi','Sanskrit','Arts','Economics','History',
                'Geography','Civis','General knowledge']
        marks=[90,89,78,95,65,84,82,81,94,87,75,41,87]
        ws.merge_cells('B{}:D{}'.format(2,2))
        i=1
        j=1
        for m in range(3, 16):
            for n in range(3,5):
                ws.merge_cells(start_row=m, start_column=2, end_row=m, end_column=3)
                border_type = 'thin'
                color = '00000000'
                ws.cell(row=m, column=n).border = Border(left=Side(border_style=border_type, color=color),
                                                         right=Side(border_style=border_type, color=color),
                                                         top=Side(border_style=border_type, color=color),
                                                         bottom=Side(border_style=border_type, color=color))
        for item in sub:

            ws.cell(row=i+2, column=2).value = item
            ws.cell(row=i+2, column=2).font = Font(color='00000000', bold=True, size=10)
            ws.cell(row=i+2, column=2).fill = PatternFill(fill_type=fills.FILL_SOLID, start_color='85C1E9',
                                                                    end_color='85C1E9')
            i+=1

        for it in marks:
            ws.cell(row=j + 2, column=4).value = it
            j+=1



    except Exception as error:
        print(error)
    wb.save(filename)
    return filename
def Marksheet():
    filename = 'SSC_marksheet.xlsx'
    wb = Workbook()
    filename = Metric_marksheet(wb, filename)


if __name__=='__main__':
    Marksheet()